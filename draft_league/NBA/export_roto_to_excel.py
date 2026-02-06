"""
导出Overall Roto数据到Excel - 用于验证计算
"""

import json
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from datetime import datetime

print("=" * 70)
print("  导出Overall Roto数据到Excel")
print("=" * 70)
print()

# 读取数据
print("读取数据...")
with open('overall_roto_rankings.json', 'r', encoding='utf-8') as f:
    rankings = json.load(f)

print(f"✅ 读取了 {len(rankings)} 个team的数据")
print()

# 创建Excel工作簿
wb = openpyxl.Workbook()

# ============================================================================
# 工作表1: Overall Rankings
# ============================================================================
ws1 = wb.active
ws1.title = "Overall Rankings"

# 设置列宽
ws1.column_dimensions['A'].width = 8   # 排名
ws1.column_dimensions['B'].width = 25  # Team
ws1.column_dimensions['C'].width = 15  # Manager
ws1.column_dimensions['D'].width = 15  # 联赛
ws1.column_dimensions['E'].width = 12  # 总积分

# 表头
headers = ['排名', 'Team', 'Manager', '联赛', '总Roto积分']
for col, header in enumerate(headers, 1):
    cell = ws1.cell(1, col, header)
    cell.font = Font(bold=True, size=11)
    cell.fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    cell.font = Font(bold=True, color="FFFFFF")
    cell.alignment = Alignment(horizontal='center', vertical='center')

# 数据
for team in rankings:
    row = ws1.max_row + 1
    ws1.cell(row, 1, team['overall_rank'])
    ws1.cell(row, 2, team['team_name'])
    ws1.cell(row, 3, team['manager'])
    ws1.cell(row, 4, team['league_name'])
    ws1.cell(row, 5, round(team['total_roto_points'], 2))
    
    # 第1名高亮
    if team['overall_rank'] == 1:
        for col in range(1, 6):
            ws1.cell(row, col).fill = PatternFill(
                start_color="FFD700", end_color="FFD700", fill_type="solid"
            )

print("✅ Overall Rankings工作表已创建")

# ============================================================================
# 工作表2: Detailed Stats (各项数据和积分)
# ============================================================================
ws2 = wb.create_sheet("Detailed Stats")

# 设置列宽
ws2.column_dimensions['A'].width = 8   # 排名
ws2.column_dimensions['B'].width = 25  # Team
ws2.column_dimensions['C'].width = 15  # 联赛
ws2.column_dimensions['D'].width = 12  # 总积分

# 11个stat，每个stat 3列（值、排名、积分）
stat_categories = ['FG%', 'FT%', '3PTM', 'PTS', 'OREB', 'REB', 'AST', 'ST', 'BLK', 'TO', 'A/T']
col_offset = 5  # 从E列开始

for i, stat in enumerate(stat_categories):
    base_col = col_offset + i * 3
    # 设置列宽
    ws2.column_dimensions[openpyxl.utils.get_column_letter(base_col)].width = 10
    ws2.column_dimensions[openpyxl.utils.get_column_letter(base_col + 1)].width = 8
    ws2.column_dimensions[openpyxl.utils.get_column_letter(base_col + 2)].width = 10

# 表头第1行（大类）
ws2.merge_cells('A1:D1')
ws2.cell(1, 1, '基本信息')
ws2.cell(1, 1).font = Font(bold=True, size=11)
ws2.cell(1, 1).fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
ws2.cell(1, 1).font = Font(bold=True, color="FFFFFF")
ws2.cell(1, 1).alignment = Alignment(horizontal='center', vertical='center')

for i, stat in enumerate(stat_categories):
    base_col = col_offset + i * 3
    ws2.merge_cells(start_row=1, start_column=base_col, end_row=1, end_column=base_col + 2)
    cell = ws2.cell(1, base_col, stat)
    cell.font = Font(bold=True, size=11)
    cell.fill = PatternFill(start_color="70AD47", end_color="70AD47", fill_type="solid")
    cell.font = Font(bold=True, color="FFFFFF")
    cell.alignment = Alignment(horizontal='center', vertical='center')

# 表头第2行（具体列）
headers_row2 = ['排名', 'Team', '联赛', '总积分']
for col, header in enumerate(headers_row2, 1):
    cell = ws2.cell(2, col, header)
    cell.font = Font(bold=True)
    cell.fill = PatternFill(start_color="B4C7E7", end_color="B4C7E7", fill_type="solid")
    cell.alignment = Alignment(horizontal='center', vertical='center')

for i, stat in enumerate(stat_categories):
    base_col = col_offset + i * 3
    ws2.cell(2, base_col, '数值').font = Font(bold=True)
    ws2.cell(2, base_col, '数值').fill = PatternFill(start_color="C6E0B4", end_color="C6E0B4", fill_type="solid")
    ws2.cell(2, base_col, '数值').alignment = Alignment(horizontal='center')
    
    ws2.cell(2, base_col + 1, '排名').font = Font(bold=True)
    ws2.cell(2, base_col + 1, '排名').fill = PatternFill(start_color="C6E0B4", end_color="C6E0B4", fill_type="solid")
    ws2.cell(2, base_col + 1, '排名').alignment = Alignment(horizontal='center')
    
    ws2.cell(2, base_col + 2, '积分').font = Font(bold=True)
    ws2.cell(2, base_col + 2, '积分').fill = PatternFill(start_color="C6E0B4", end_color="C6E0B4", fill_type="solid")
    ws2.cell(2, base_col + 2, '积分').alignment = Alignment(horizontal='center')

# 数据行
for team in rankings:
    row = ws2.max_row + 1
    
    # 基本信息
    ws2.cell(row, 1, team['overall_rank'])
    ws2.cell(row, 2, team['team_name'])
    ws2.cell(row, 3, team['league_name'])
    ws2.cell(row, 4, round(team['total_roto_points'], 2))
    
    # 各项stat
    stats = team['stats']
    for i, stat in enumerate(stat_categories):
        base_col = col_offset + i * 3
        stat_info = stats.get(stat, {})
        
        # 数值
        value = stat_info.get('value', 0)
        ws2.cell(row, base_col, round(value, 3))
        
        # 排名
        rank = stat_info.get('rank', 0)
        ws2.cell(row, base_col + 1, rank)
        
        # 积分
        points = stat_info.get('roto_points', 0)
        ws2.cell(row, base_col + 2, round(points, 2))
        
        # 排名靠前的高亮（前10名）
        if rank <= 10:
            ws2.cell(row, base_col + 1).fill = PatternFill(
                start_color="92D050", end_color="92D050", fill_type="solid"
            )
        # 排名靠后的标注（后10名）
        elif rank >= 183:
            ws2.cell(row, base_col + 1).fill = PatternFill(
                start_color="FFC7CE", end_color="FFC7CE", fill_type="solid"
            )

print("✅ Detailed Stats工作表已创建")

# ============================================================================
# 工作表3: Stats by Category (每个stat单独排名)
# ============================================================================
ws3 = wb.create_sheet("Stats by Category")

# 为每个stat创建排名
current_row = 1

for stat in stat_categories:
    # 标题
    ws3.merge_cells(start_row=current_row, start_column=1, end_row=current_row, end_column=5)
    title_cell = ws3.cell(current_row, 1, f"{stat} Rankings")
    title_cell.font = Font(bold=True, size=12)
    title_cell.fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    title_cell.font = Font(bold=True, color="FFFFFF")
    title_cell.alignment = Alignment(horizontal='center')
    current_row += 1
    
    # 表头
    headers = ['排名', 'Team', '联赛', '数值', 'Roto积分']
    for col, header in enumerate(headers, 1):
        cell = ws3.cell(current_row, col, header)
        cell.font = Font(bold=True)
        cell.fill = PatternFill(start_color="B4C7E7", end_color="B4C7E7", fill_type="solid")
        cell.alignment = Alignment(horizontal='center')
    current_row += 1
    
    # 获取这个stat的所有team并排序
    stat_data = []
    for team in rankings:
        stat_info = team['stats'].get(stat, {})
        stat_data.append({
            'team_name': team['team_name'],
            'league_name': team['league_name'],
            'value': stat_info.get('value', 0),
            'rank': stat_info.get('rank', 0),
            'roto_points': stat_info.get('roto_points', 0)
        })
    
    # 按rank排序
    stat_data.sort(key=lambda x: x['rank'])
    
    # 写入数据（只显示前20）
    for i, data in enumerate(stat_data[:20], 1):
        ws3.cell(current_row, 1, data['rank'])
        ws3.cell(current_row, 2, data['team_name'])
        ws3.cell(current_row, 3, data['league_name'])
        ws3.cell(current_row, 4, round(data['value'], 3))
        ws3.cell(current_row, 5, round(data['roto_points'], 2))
        
        # 前3名高亮
        if i <= 3:
            fill_color = "FFD700" if i == 1 else ("C0C0C0" if i == 2 else "CD7F32")
            for col in range(1, 6):
                ws3.cell(current_row, col).fill = PatternFill(
                    start_color=fill_color, end_color=fill_color, fill_type="solid"
                )
        
        current_row += 1
    
    current_row += 2  # 空行

print("✅ Stats by Category工作表已创建")

# ============================================================================
# 工作表4: League Summary (各联赛统计)
# ============================================================================
ws4 = wb.create_sheet("League Summary")

# 统计各联赛
league_stats = {}
for team in rankings:
    league = team['league_name']
    if league not in league_stats:
        league_stats[league] = {
            'count': 0,
            'total_points': 0,
            'max_points': 0,
            'min_points': float('inf'),
            'best_team': '',
            'best_rank': 999
        }
    
    league_stats[league]['count'] += 1
    league_stats[league]['total_points'] += team['total_roto_points']
    
    if team['total_roto_points'] > league_stats[league]['max_points']:
        league_stats[league]['max_points'] = team['total_roto_points']
        league_stats[league]['best_team'] = team['team_name']
        league_stats[league]['best_rank'] = team['overall_rank']
    
    if team['total_roto_points'] < league_stats[league]['min_points']:
        league_stats[league]['min_points'] = team['total_roto_points']

# 表头
headers = ['联赛', 'Team数', '平均积分', '最高积分', '最低积分', '最佳Team', '总排名']
for col, header in enumerate(headers, 1):
    cell = ws4.cell(1, col, header)
    cell.font = Font(bold=True)
    cell.fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    cell.font = Font(bold=True, color="FFFFFF")
    cell.alignment = Alignment(horizontal='center')

# 数据
row = 2
for league, stats in sorted(league_stats.items()):
    ws4.cell(row, 1, league)
    ws4.cell(row, 2, stats['count'])
    ws4.cell(row, 3, round(stats['total_points'] / stats['count'], 2))
    ws4.cell(row, 4, round(stats['max_points'], 2))
    ws4.cell(row, 5, round(stats['min_points'], 2))
    ws4.cell(row, 6, stats['best_team'])
    ws4.cell(row, 7, stats['best_rank'])
    row += 1

# 设置列宽
ws4.column_dimensions['A'].width = 15
ws4.column_dimensions['B'].width = 10
ws4.column_dimensions['C'].width = 12
ws4.column_dimensions['D'].width = 12
ws4.column_dimensions['E'].width = 12
ws4.column_dimensions['F'].width = 25
ws4.column_dimensions['G'].width = 10

print("✅ League Summary工作表已创建")

# 保存文件
filename = f'overall_roto_rankings_{datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsx'
wb.save(filename)

print()
print("=" * 70)
print("  ✅ 导出完成！")
print("=" * 70)
print()
print(f"文件名: {filename}")
print()
print("包含的工作表:")
print("  1. Overall Rankings - 总排名")
print("  2. Detailed Stats - 详细stats（值、排名、积分）")
print("  3. Stats by Category - 各stat单独排名（Top 20）")
print("  4. League Summary - 各联赛统计")
print()
print("现在可以打开Excel文件验证计算结果！")
print()
