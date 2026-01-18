#!/usr/bin/env python3
"""
铁人团队赛 - 初始化数据库并导入数据
从Excel导入团队信息和已完成项目的得分
"""

import sqlite3
import openpyxl
import sys
import os
from datetime import datetime

def create_database(db_path):
    """创建数据库结构"""
    print("="*60)
    print("创建数据库结构")
    print("="*60)
    
    # 读取SQL文件
    sql_file = os.path.join(os.path.dirname(__file__), 'irongroup_database.sql')
    with open(sql_file, 'r', encoding='utf-8') as f:
        sql_script = f.read()
    
    conn = sqlite3.connect(db_path)
    cursor = conn.cursor()
    
    # 执行SQL脚本
    cursor.executescript(sql_script)
    conn.commit()
    conn.close()
    
    print("✓ 数据库结构创建完成\n")

def import_teams_and_members(excel_path, db_path):
    """从Excel导入团队和成员信息"""
    print("="*60)
    print("导入团队和成员信息")
    print("="*60)
    
    wb = openpyxl.load_workbook(excel_path, data_only=True)
    sheet = wb['组队赛花名册']
    
    conn = sqlite3.connect(db_path)
    cursor = conn.cursor()
    
    # 导入12个团队
    for row_idx in range(2, 14):  # 第2行到第13行
        team_name = sheet.cell(row_idx, 1).value
        members_str = sheet.cell(row_idx, 2).value
        
        if not team_name or not members_str:
            continue
        
        # 插入团队
        cursor.execute('INSERT OR IGNORE INTO teams (team_name) VALUES (?)', (team_name,))
        
        # 获取team_id
        cursor.execute('SELECT team_id FROM teams WHERE team_name = ?', (team_name,))
        team_id = cursor.fetchone()[0]
        
        # 解析成员（用中文顿号或逗号分隔）
        members = [m.strip() for m in members_str.replace('、', ',').split(',')]
        
        # 插入成员
        for member in members:
            if member:
                cursor.execute(
                    'INSERT OR IGNORE INTO team_members (team_id, member_name) VALUES (?, ?)',
                    (team_id, member)
                )
        
        print(f"✓ {team_name}: {', '.join(members)}")
    
    conn.commit()
    conn.close()
    print(f"\n✓ 成功导入 12 个团队\n")

def import_leagues(db_path):
    """初始化联赛信息"""
    print("="*60)
    print("初始化联赛信息")
    print("="*60)
    
    conn = sqlite3.connect(db_path)
    cursor = conn.cursor()
    
    leagues = [
        ('team_mlb', 'MLB', 'Yahoo', 'completed', None, None),
        ('team_nfl', 'NFL', 'Yahoo', 'completed', 'nfl.l.749886', None),
        ('team_nhl', 'NHL', 'Yahoo', 'active', 'nhl.l.29114', None),
        ('team_nba', 'NBA', 'Yahoo', 'active', 'nba.l.84043', None),
        ('team_epl', 'EPL', 'Fantrax', 'active', None, 'w8kxeqp2mcnclqc9'),
    ]
    
    for league_id, sport, platform, status, yahoo_id, fantrax_id in leagues:
        cursor.execute('''
            INSERT OR REPLACE INTO leagues 
            (league_id, sport, platform, status, yahoo_league_id, fantrax_league_id, updated_at)
            VALUES (?, ?, ?, ?, ?, ?, datetime('now'))
        ''', (league_id, sport, platform, status, yahoo_id, fantrax_id))
        
        status_icon = "✅" if status == 'completed' else "🔄"
        print(f"  {status_icon} {sport}: {platform} (League: {yahoo_id or fantrax_id})")
    
    conn.commit()
    conn.close()
    print("\n✓ 联赛信息初始化完成\n")

def import_completed_scores(excel_path, db_path):
    """从Excel导入已完成项目的得分（MLB和NFL）"""
    print("="*60)
    print("导入已完成项目得分（MLB和NFL）")
    print("="*60)
    
    wb = openpyxl.load_workbook(excel_path, data_only=True)
    sheet = wb['组队赛计分']
    
    conn = sqlite3.connect(db_path)
    cursor = conn.cursor()
    
    # 遍历所有球队（第4行开始）
    for row_idx in range(4, 16):
        team_name = sheet.cell(row_idx, 5).value  # E列：球队名
        if not team_name:
            break
        
        # 获取team_id
        cursor.execute('SELECT team_id FROM teams WHERE team_name = ?', (team_name,))
        result = cursor.fetchone()
        if not result:
            print(f"⚠ 警告：找不到团队 {team_name}")
            continue
        team_id = result[0]
        
        print(f"\n{team_name}:")
        
        # === MLB得分（已完成）===
        mlb_po_rank = sheet.cell(row_idx, 8).value   # H列：季后赛排名
        mlb_score = sheet.cell(row_idx, 9).value     # I列：计分
        mlb_regular = sheet.cell(row_idx, 10).value  # J列：常规赛加分
        
        if mlb_score:
            mlb_total = mlb_score + (mlb_regular or 0)
            cursor.execute('''
                INSERT OR REPLACE INTO team_scores 
                (team_id, league_id, sport, playoff_rank, playoff_score, 
                 regular_bonus, total_score, is_final, updated_at)
                VALUES (?, ?, ?, ?, ?, ?, ?, ?, datetime('now'))
            ''', (team_id, 'team_mlb', 'MLB', mlb_po_rank, mlb_score, 
                  mlb_regular or 0, mlb_total, 1))
            print(f"  MLB: 季后赛第{mlb_po_rank}名 = {mlb_score}分 + 常规赛{mlb_regular or 0}分 = {mlb_total}分")
        
        # === NFL得分（已完成）===
        nfl_po_rank = sheet.cell(row_idx, 11).value  # K列：季后赛排名
        nfl_score = sheet.cell(row_idx, 12).value    # L列：计分
        nfl_regular = sheet.cell(row_idx, 13).value  # M列：常规赛加分
        
        if nfl_score:
            nfl_total = nfl_score + (nfl_regular or 0)
            cursor.execute('''
                INSERT OR REPLACE INTO team_scores 
                (team_id, league_id, sport, playoff_rank, playoff_score, 
                 regular_bonus, total_score, is_final, updated_at)
                VALUES (?, ?, ?, ?, ?, ?, ?, ?, datetime('now'))
            ''', (team_id, 'team_nfl', 'NFL', nfl_po_rank, nfl_score, 
                  nfl_regular or 0, nfl_total, 1))
            print(f"  NFL: 季后赛第{nfl_po_rank}名 = {nfl_score}分 + 常规赛{nfl_regular or 0}分 = {nfl_total}分")
    
    conn.commit()
    conn.close()
    print("\n✓ MLB和NFL得分导入完成\n")

def update_leaderboard(db_path):
    """更新团队排行榜"""
    print("="*60)
    print("更新团队排行榜")
    print("="*60)
    
    conn = sqlite3.connect(db_path)
    cursor = conn.cursor()
    
    # 计算每个团队的总分
    cursor.execute('''
        SELECT 
            t.team_id,
            t.team_name,
            COALESCE(SUM(s.total_score), 0) as total_score,
            COALESCE(MAX(CASE WHEN s.sport = 'MLB' THEN s.total_score END), 0) as mlb,
            COALESCE(MAX(CASE WHEN s.sport = 'NFL' THEN s.total_score END), 0) as nfl,
            COALESCE(MAX(CASE WHEN s.sport = 'NHL' THEN s.total_score END), 0) as nhl,
            COALESCE(MAX(CASE WHEN s.sport = 'NBA' THEN s.total_score END), 0) as nba,
            COALESCE(MAX(CASE WHEN s.sport = 'EPL' THEN s.total_score END), 0) as epl,
            COUNT(CASE WHEN s.is_final = 1 THEN 1 END) as completed
        FROM teams t
        LEFT JOIN team_scores s ON t.team_id = s.team_id
        GROUP BY t.team_id, t.team_name
        ORDER BY total_score DESC, t.team_name
    ''')
    
    teams = cursor.fetchall()
    
    # 清空并重新插入排行榜
    cursor.execute('DELETE FROM team_leaderboard')
    
    print("\n排名  球队                         总分   MLB   NFL   完成")
    print("-" * 65)
    
    for rank, (team_id, team_name, total, mlb, nfl, nhl, nba, epl, completed) in enumerate(teams, 1):
        cursor.execute('''
            INSERT INTO team_leaderboard
            (team_id, team_name, rank, total_score, 
             mlb_score, nfl_score, nhl_score, nba_score, epl_score,
             completed_sports, updated_at)
            VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, datetime('now'))
        ''', (team_id, team_name, rank, total, mlb, nfl, nhl, nba, epl, completed))
        
        print(f"{rank:2d}.   {team_name:28s} {total:5.1f}  {mlb:4.1f}  {nfl:4.1f}  {completed}/5")
    
    conn.commit()
    conn.close()
    print("\n✓ 排行榜更新完成\n")

def main():
    if len(sys.argv) < 2:
        print("用法: python import_irongroup_data.py <Excel文件路径> [数据库路径]")
        print("示例: python import_irongroup_data.py 2526铁人团队赛.xlsx irongroup.db")
        sys.exit(1)
    
    excel_path = sys.argv[1]
    db_path = sys.argv[2] if len(sys.argv) > 2 else 'irongroup.db'
    
    if not os.path.exists(excel_path):
        print(f"错误：找不到Excel文件 {excel_path}")
        sys.exit(1)
    
    print("\n" + "="*60)
    print("铁人团队赛 - 数据导入")
    print("="*60)
    print(f"Excel文件: {excel_path}")
    print(f"数据库: {db_path}")
    print("="*60 + "\n")
    
    # 执行导入步骤
    create_database(db_path)
    import_teams_and_members(excel_path, db_path)
    import_leagues(db_path)
    import_completed_scores(excel_path, db_path)
    update_leaderboard(db_path)
    
    print("="*60)
    print("✓ 数据导入完成！")
    print("="*60)
    print(f"\n数据库文件: {db_path}")
    print(f"团队数量: 12")
    print(f"已完成项目: 2 (MLB, NFL)")
    print(f"进行中项目: 3 (NHL, NBA, EPL)")
    print("\n下一步:")
    print("  1. 运行Flask应用: python irongroup_app.py")
    print("  2. 访问: http://127.0.0.1:5000/irongroup/leaderboard")
    print()

if __name__ == '__main__':
    main()
