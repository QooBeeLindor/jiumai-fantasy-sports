#!/usr/bin/env python3
"""
根据实际选秀结果更新数据库中的队名
使用选秀顺位来匹配玩家
"""

import sqlite3
import openpyxl

def update_team_names_from_draft(excel_path, db_path):
    """
    从Excel的选秀结果更新数据库队名
    """
    print("="*60)
    print("根据选秀结果更新队名")
    print("="*60)
    
    conn = sqlite3.connect(db_path)
    cursor = conn.cursor()
    
    wb = openpyxl.load_workbook(excel_path)
    
    # 1. 读取选秀顺位对照表（哪个玩家在哪个项目第几顺位）
    draft_order_sheet = wb['选秀顺位']
    
    # 建立映射：{sport: {顺位: 玩家名}}
    sport_draft_map = {
        'MLB': {},
        'NHL': {},
        'NBA': {},
        'NFL': {}
    }
    
    print("\n[1/3] 读取选秀顺位对照表...")
    for row in draft_order_sheet.iter_rows(min_row=2, max_row=17, values_only=True):
        pick = row[1]  # B列：顺位
        if pick:
            sport_draft_map['MLB'][int(pick)] = row[2]  # C列：MLB玩家
            sport_draft_map['NHL'][int(pick)] = row[3]  # D列：NHL玩家
            sport_draft_map['NBA'][int(pick)] = row[4]  # E列：NBA玩家
            sport_draft_map['NFL'][int(pick)] = row[5]  # F列：NFL玩家
    
    print(f"  ✓ MLB: {len(sport_draft_map['MLB'])}个顺位")
    print(f"  ✓ NHL: {len(sport_draft_map['NHL'])}个顺位")
    print(f"  ✓ NBA: {len(sport_draft_map['NBA'])}个顺位")
    print(f"  ✓ NFL: {len(sport_draft_map['NFL'])}个顺位")
    
    # 2. 读取实际选秀结果（Yahoo队名）
    updates = []
    
    print("\n[2/4] 读取实际选秀结果...")
    
    # MLB - 设为玩家名本身（找不到选秀结果）
    print("\n  MLB: 使用玩家名作为队名")
    for pick, player_name in sport_draft_map['MLB'].items():
        updates.append(('MLB', player_name, player_name))
        print(f"    {pick:2d}. {player_name:15s} → {player_name}")
    
    # NFL选秀结果
    if 'NFL项选秀首轮' in wb.sheetnames:
        sheet = wb['NFL项选秀首轮']
        print("\n  NFL选秀结果:")
        for row in sheet.iter_rows(min_row=2, max_row=17, values_only=True):
            if row[1]:  # B列：有顺位
                pick = int(row[1])
                yahoo_team_name = row[3]  # D列：玩家队名
                if yahoo_team_name and pick in sport_draft_map['NFL']:
                    player_name = sport_draft_map['NFL'][pick]
                    updates.append(('NFL', player_name, yahoo_team_name))
                    print(f"    {pick:2d}. {player_name:15s} → {yahoo_team_name}")
    
    # NHL选秀结果
    if 'NHL项选秀首轮' in wb.sheetnames:
        sheet = wb['NHL项选秀首轮']
        print("\n  NHL选秀结果:")
        for row in sheet.iter_rows(min_row=2, max_row=17, values_only=True):
            if row[1]:  # B列：有顺位
                pick = int(row[1])
                yahoo_team_name = row[3]  # D列：玩家队名
                if yahoo_team_name and pick in sport_draft_map['NHL']:
                    player_name = sport_draft_map['NHL'][pick]
                    updates.append(('NHL', player_name, yahoo_team_name))
                    print(f"    {pick:2d}. {player_name:15s} → {yahoo_team_name}")
    
    # NBA选秀结果
    if 'NBA项选秀首轮' in wb.sheetnames:
        sheet = wb['NBA项选秀首轮']
        print("\n  NBA选秀结果:")
        for row in sheet.iter_rows(min_row=2, max_row=17, values_only=True):
            if row[1]:  # B列：有顺位
                pick = int(row[1])
                yahoo_team_name = row[3]  # D列：玩家队名
                if yahoo_team_name and pick in sport_draft_map['NBA']:
                    player_name = sport_draft_map['NBA'][pick]
                    updates.append(('NBA', player_name, yahoo_team_name))
                    print(f"    {pick:2d}. {player_name:15s} → {yahoo_team_name}")
    
    # 3. 更新数据库
    print(f"\n[4/4] 更新数据库...")
    print(f"  共需要更新 {len(updates)} 条记录\n")
    
    updated_count = 0
    for sport, player_name, yahoo_team_name in updates:
        # 获取player_id
        cursor.execute('SELECT player_id FROM players WHERE player_name = ?', 
                      (player_name,))
        result = cursor.fetchone()
        
        if result:
            player_id = result[0]
            
            # 更新或插入队名
            cursor.execute('''
                INSERT OR REPLACE INTO sport_mappings 
                (player_id, sport, yahoo_team_name)
                VALUES (?, ?, ?)
            ''', (player_id, sport, yahoo_team_name))
            
            print(f"  ✓ {sport:4s} | {player_name:15s} → {yahoo_team_name}")
            updated_count += 1
        else:
            print(f"  ⚠ 找不到玩家: {player_name}")
    
    conn.commit()
    conn.close()
    
    print("\n" + "="*60)
    print(f"✓ 完成！成功更新 {updated_count} 条队名记录")
    print("="*60)
    print("\n队名设置说明:")
    print("  - MLB: 使用玩家名作为队名（找不到选秀结果）")
    print("  - NFL: 从选秀结果获取实际队名")
    print("  - NHL: 从选秀结果获取实际队名")
    print("  - NBA: 从选秀结果获取实际队名")
    print("\n下一步:")
    print("  1. 运行 check_team_mappings.py 查看更新后的队名")
    print("  2. 运行 sync_yahoo_standings.py 同步数据")


def main():
    import sys
    
    excel_path = '/mnt/project/2526铁人个人赛.xlsx' if len(sys.argv) < 2 else sys.argv[1]
    db_path = 'ironman.db' if len(sys.argv) < 3 else sys.argv[2]
    
    update_team_names_from_draft(excel_path, db_path)


if __name__ == '__main__':
    main()
