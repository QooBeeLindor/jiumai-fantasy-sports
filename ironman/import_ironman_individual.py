#!/usr/bin/env python3
"""
铁人个人赛 - Excel数据导入工具
功能：导入玩家信息、队名映射、联赛信息
版本：v2.0
日期：2026-01-11
"""

import sqlite3
import openpyxl
from datetime import datetime

class IronmanIndividualImporter:
    def __init__(self, excel_path, db_path):
        self.excel_path = excel_path
        self.db_path = db_path
        self.conn = None
        self.cursor = None
        
    def connect_db(self):
        """连接数据库"""
        self.conn = sqlite3.connect(self.db_path)
        self.cursor = self.conn.cursor()
        print(f"✓ 已连接到数据库: {self.db_path}")
        
    def close_db(self):
        """关闭数据库连接"""
        if self.conn:
            self.conn.commit()
            self.conn.close()
            print("✓ 数据库连接已关闭")
    
    def import_leagues(self):
        """导入联赛信息"""
        print("\n=== 导入联赛信息 ===")
        
        wb = openpyxl.load_workbook(self.excel_path)
        sheet = wb['个人赛赛各盟链接']
        
        leagues = []
        for row in sheet.iter_rows(min_row=2, max_row=5, values_only=True):
            if row[0] and row[1]:
                sport = row[0]
                link = row[1]
                
                # 从链接中提取League ID
                league_id = None
                if 'fantasysports.yahoo.com' in link:
                    # 格式: .../nba/68792/ 或 .../f1/745126
                    parts = link.rstrip('/').split('/')
                    league_id = parts[-1]
                
                # 判断状态（MLB已关闭）
                status = 'completed' if 'MLB' in sport or '已关闭' in link else 'active'
                
                if league_id:
                    leagues.append((league_id, sport, f"{sport}铁人个人赛", '2024-25', status))
                    print(f"  ✓ {sport:5s}: League ID={league_id} (状态: {status})")
        
        # 插入数据库
        for league_id, sport, name, season, status in leagues:
            self.cursor.execute('''
                INSERT OR REPLACE INTO leagues 
                (league_id, sport, league_name, season, status)
                VALUES (?, ?, ?, ?, ?)
            ''', (league_id, sport, name, season, status))
        
        self.conn.commit()
        print(f"  共导入 {len(leagues)} 个联赛")
    
    def import_players_and_mappings(self):
        """导入玩家信息和队名映射"""
        print("\n=== 导入玩家和队名映射 ===")
        
        wb = openpyxl.load_workbook(self.excel_path)
        sheet = wb['个人赛花名册']
        
        players = []
        mappings = []
        
        # 读取玩家数据
        for row in sheet.iter_rows(min_row=2, max_row=17, values_only=True):
            if row[3]:  # 铁人个人赛玩家名
                draft_order = row[2]
                player_name = row[3]
                
                # MLB、NHL、NBA、NFL的队名
                mlb_name = row[6] if len(row) > 6 else None
                nhl_name = row[7] if len(row) > 7 else None
                nba_name = row[8] if len(row) > 8 else None
                nfl_name = row[9] if len(row) > 9 else None
                
                players.append((player_name, draft_order))
                
                # 记录队名映射
                mappings.append((player_name, 'MLB', mlb_name))
                mappings.append((player_name, 'NHL', nhl_name))
                mappings.append((player_name, 'NBA', nba_name))
                mappings.append((player_name, 'NFL', nfl_name))
        
        # 插入玩家
        player_id_map = {}
        for player_name, draft_order in players:
            self.cursor.execute('''
                INSERT OR IGNORE INTO players (player_name, draft_order)
                VALUES (?, ?)
            ''', (player_name, draft_order))
            
            # 获取player_id
            self.cursor.execute('SELECT player_id FROM players WHERE player_name = ?', 
                              (player_name,))
            player_id = self.cursor.fetchone()[0]
            player_id_map[player_name] = player_id
        
        print(f"  ✓ 导入玩家: {len(players)}名")
        
        # 插入队名映射
        mapping_count = 0
        for player_name, sport, yahoo_team_name in mappings:
            if yahoo_team_name:  # 只插入有队名的
                player_id = player_id_map[player_name]
                self.cursor.execute('''
                    INSERT OR REPLACE INTO sport_mappings 
                    (player_id, sport, yahoo_team_name)
                    VALUES (?, ?, ?)
                ''', (player_id, sport, yahoo_team_name))
                mapping_count += 1
        
        self.conn.commit()
        print(f"  ✓ 导入队名映射: {mapping_count}条")
        
        # 显示玩家和映射示例
        print("\n  玩家列表（前5名）:")
        for i, (name, order) in enumerate(players[:5], 1):
            print(f"    {order:2d}. {name}")
        
        print("\n  队名映射示例（GuSone）:")
        if 'GuSone' in player_id_map:
            self.cursor.execute('''
                SELECT sport, yahoo_team_name 
                FROM sport_mappings 
                WHERE player_id = ?
                ORDER BY sport
            ''', (player_id_map['GuSone'],))
            for sport, team_name in self.cursor.fetchall():
                print(f"    {sport}: {team_name}")
    
    def import_completed_scores(self):
        """导入已结束项目的积分（MLB、NFL）"""
        print("\n=== 导入已结束项目积分 ===")
        
        wb = openpyxl.load_workbook(self.excel_path)
        sheet = wb['个人赛计分 ']  # 注意：sheet名有尾随空格
        
        # 读取已有积分数据
        scores_data = []
        for row in sheet.iter_rows(min_row=4, max_row=19, values_only=True):
            player_name = row[4] if len(row) > 4 else None
            if player_name:
                # MLB常规赛排名和季后赛积分（列H-J）
                mlb_regular_rank = row[7] if len(row) > 7 and row[7] else None
                mlb_playoff_points = row[9] if len(row) > 9 and row[9] else 0
                
                # NFL常规赛排名和季后赛积分（列K-M）
                nfl_regular_rank = row[10] if len(row) > 10 and row[10] else None
                nfl_playoff_points = row[12] if len(row) > 12 and row[12] else 0
                
                scores_data.append({
                    'player': player_name,
                    'mlb_rank': mlb_regular_rank,
                    'mlb_playoff': mlb_playoff_points,
                    'nfl_rank': nfl_regular_rank,
                    'nfl_playoff': nfl_playoff_points
                })
        
        # 获取player_id映射
        self.cursor.execute('SELECT player_id, player_name FROM players')
        player_map = {name: pid for pid, name in self.cursor.fetchall()}
        
        # 获取积分规则
        self.cursor.execute('SELECT rank_position, points FROM scoring_rules WHERE stage="regular"')
        regular_points_map = {rank: points for rank, points in self.cursor.fetchall()}
        
        # 插入积分数据
        inserted = 0
        for data in scores_data:
            if data['player'] not in player_map:
                continue
            
            player_id = player_map[data['player']]
            
            # MLB积分
            if data['mlb_rank']:
                mlb_regular_points = regular_points_map.get(int(data['mlb_rank']), 0)
                mlb_playoff_points = float(data['mlb_playoff']) if data['mlb_playoff'] else 0
                mlb_total = mlb_regular_points + mlb_playoff_points
                
                self.cursor.execute('''
                    INSERT OR REPLACE INTO ironman_scores
                    (player_id, sport, regular_rank, regular_points, 
                     playoff_points, total_points, is_final)
                    VALUES (?, ?, ?, ?, ?, ?, ?)
                ''', (player_id, 'MLB', int(data['mlb_rank']), mlb_regular_points,
                      mlb_playoff_points, mlb_total, 1))
                inserted += 1
            
            # NFL积分
            if data['nfl_rank']:
                nfl_regular_points = regular_points_map.get(int(data['nfl_rank']), 0)
                nfl_playoff_points = float(data['nfl_playoff']) if data['nfl_playoff'] else 0
                nfl_total = nfl_regular_points + nfl_playoff_points
                
                self.cursor.execute('''
                    INSERT OR REPLACE INTO ironman_scores
                    (player_id, sport, regular_rank, regular_points,
                     playoff_points, total_points, is_final)
                    VALUES (?, ?, ?, ?, ?, ?, ?)
                ''', (player_id, 'NFL', int(data['nfl_rank']), nfl_regular_points,
                      nfl_playoff_points, nfl_total, 1))
                inserted += 1
        
        self.conn.commit()
        print(f"  ✓ 导入已结束项目积分: {inserted}条记录")
    
    def update_leaderboard(self):
        """更新总排行榜"""
        print("\n=== 更新总排行榜 ===")
        
        # 查询所有玩家及其积分
        self.cursor.execute('''
            SELECT 
                p.player_id,
                p.player_name,
                COALESCE(SUM(s.total_points), 0) as total_score,
                MAX(CASE WHEN s.sport='MLB' THEN s.total_points ELSE 0 END) as mlb_points,
                MAX(CASE WHEN s.sport='NFL' THEN s.total_points ELSE 0 END) as nfl_points,
                MAX(CASE WHEN s.sport='NHL' THEN s.total_points ELSE 0 END) as nhl_points,
                MAX(CASE WHEN s.sport='NBA' THEN s.total_points ELSE 0 END) as nba_points,
                COUNT(CASE WHEN s.is_final=1 THEN 1 END) as completed_sports
            FROM players p
            LEFT JOIN ironman_scores s ON p.player_id = s.player_id
            GROUP BY p.player_id, p.player_name
        ''')
        
        leaderboard = self.cursor.fetchall()
        
        # 插入或更新排行榜
        for row in leaderboard:
            player_id, player_name, total_score, mlb, nfl, nhl, nba, completed = row
            
            self.cursor.execute('''
                INSERT OR REPLACE INTO ironman_leaderboard
                (player_id, player_name, total_score, mlb_points, nfl_points,
                 nhl_points, nba_points, completed_sports)
                VALUES (?, ?, ?, ?, ?, ?, ?, ?)
            ''', (player_id, player_name, total_score, mlb, nfl, nhl, nba, completed))
        
        # 更新排名
        self.cursor.execute('''
            UPDATE ironman_leaderboard
            SET rank = (
                SELECT COUNT(*) + 1
                FROM ironman_leaderboard l2
                WHERE l2.total_score > ironman_leaderboard.total_score
            )
        ''')
        
        self.conn.commit()
        print(f"  ✓ 更新排行榜: {len(leaderboard)}名玩家")
    
    def show_results(self):
        """显示导入结果"""
        print("\n" + "="*60)
        print("数据导入结果")
        print("="*60)
        
        # 显示TOP5排行
        print("\n【铁人个人赛TOP5】")
        self.cursor.execute('''
            SELECT rank, player_name, total_score, mlb_points, nfl_points,
                   completed_sports
            FROM ironman_leaderboard
            ORDER BY rank
            LIMIT 5
        ''')
        
        print(f"{'排名':<6} {'玩家':<12} {'总分':<8} {'MLB':<6} {'NFL':<6} {'完成项目'}")
        print("-" * 60)
        for rank, name, total, mlb, nfl, completed in self.cursor.fetchall():
            print(f"{rank:<6} {name:<12} {total:<8.1f} {mlb:<6.1f} {nfl:<6.1f} {completed}/4")
        
        # 统计信息
        print("\n【统计信息】")
        self.cursor.execute('SELECT COUNT(*) FROM players')
        print(f"  玩家总数: {self.cursor.fetchone()[0]}")
        
        self.cursor.execute('SELECT COUNT(*) FROM sport_mappings')
        print(f"  队名映射: {self.cursor.fetchone()[0]}")
        
        self.cursor.execute('SELECT COUNT(*) FROM leagues')
        print(f"  联赛数量: {self.cursor.fetchone()[0]}")
        
        self.cursor.execute('SELECT COUNT(*) FROM ironman_scores WHERE is_final=1')
        print(f"  已完成项目积分: {self.cursor.fetchone()[0]}")
    
    def run(self):
        """执行完整导入流程"""
        print("="*60)
        print("铁人个人赛 - 数据导入工具")
        print("="*60)
        
        try:
            self.connect_db()
            
            # 依次导入
            self.import_leagues()
            self.import_players_and_mappings()
            self.import_completed_scores()
            self.update_leaderboard()
            
            # 显示结果
            self.show_results()
            
            print("\n" + "="*60)
            print("✓ 数据导入完成！")
            print("="*60)
            
        except Exception as e:
            print(f"\n✗ 导入失败: {str(e)}")
            import traceback
            traceback.print_exc()
            
        finally:
            self.close_db()


def main():
    import sys
    
    # 默认路径
    excel_path = '/mnt/project/2526铁人个人赛.xlsx'
    db_path = 'ironman.db'
    
    # 支持命令行参数
    if len(sys.argv) > 1:
        excel_path = sys.argv[1]
    if len(sys.argv) > 2:
        db_path = sys.argv[2]
    
    print(f"Excel文件: {excel_path}")
    print(f"数据库文件: {db_path}\n")
    
    importer = IronmanIndividualImporter(excel_path, db_path)
    importer.run()


if __name__ == '__main__':
    main()
