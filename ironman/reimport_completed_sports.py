#!/usr/bin/env python3
"""
从Excel重新导入已结束项目的积分数据
适用于：MLB、NFL等已结束的项目
"""

import sqlite3
import openpyxl
import sys

def reimport_completed_sports(excel_path, db_path):
    """
    从Excel"个人赛计分"sheet重新导入已结束项目的数据
    """
    print("="*60)
    print("重新导入已结束项目数据")
    print("="*60)
    print(f"\nExcel: {excel_path}")
    print(f"数据库: {db_path}\n")
    
    # 连接数据库
    conn = sqlite3.connect(db_path)
    cursor = conn.cursor()
    print("✓ 已连接到数据库")
    
    # 打开Excel（data_only=True以获取公式计算后的值）
    wb = openpyxl.load_workbook(excel_path, data_only=True)
    sheet = wb['个人赛计分 ']
    print("✓ 已打开Excel计分表\n")
    
    print("="*60)
    print("[1/3] 读取Excel数据")
    print("="*60 + "\n")
    
    # 读取玩家数据
    updates = []
    for row_idx in range(4, 20):  # 第4-19行是玩家数据
        player_name = sheet.cell(row_idx, 5).value  # E列：玩家名
        if not player_name:
            continue
        
        # 获取player_id
        cursor.execute('SELECT player_id FROM players WHERE player_name = ?', (player_name,))
        result = cursor.fetchone()
        if not result:
            print(f"  ⚠ 找不到玩家: {player_name}")
            continue
        player_id = result[0]
        
        # MLB数据 (H, I, J列)
        mlb_rank = sheet.cell(row_idx, 8).value
        mlb_regular = sheet.cell(row_idx, 9).value
        mlb_playoff = sheet.cell(row_idx, 10).value
        
        if mlb_rank and mlb_regular is not None:
            mlb_total = float(mlb_regular or 0) + float(mlb_playoff or 0)
            updates.append({
                'player_id': player_id,
                'player_name': player_name,
                'sport': 'MLB',
                'regular_rank': int(float(mlb_rank)) if mlb_rank else None,
                'regular_points': float(mlb_regular or 0),
                'playoff_points': float(mlb_playoff or 0),
                'total_points': mlb_total
            })
            print(f"  {player_name:15s} MLB: 常规#{int(float(mlb_rank)):2d} ({mlb_regular:5.1f}) + 季后({mlb_playoff or 0:4.1f}) = {mlb_total:5.1f}")
        
        # NFL数据 (K, L, M列)
        nfl_rank = sheet.cell(row_idx, 11).value
        nfl_regular = sheet.cell(row_idx, 12).value
        nfl_playoff = sheet.cell(row_idx, 13).value
        
        if nfl_rank and nfl_regular is not None:
            nfl_total = float(nfl_regular or 0) + float(nfl_playoff or 0)
            updates.append({
                'player_id': player_id,
                'player_name': player_name,
                'sport': 'NFL',
                'regular_rank': int(float(nfl_rank)) if nfl_rank else None,
                'regular_points': float(nfl_regular or 0),
                'playoff_points': float(nfl_playoff or 0),
                'total_points': nfl_total
            })
            print(f"  {player_name:15s} NFL: 常规#{int(float(nfl_rank)):2d} ({nfl_regular:5.1f}) + 季后({nfl_playoff or 0:4.1f}) = {nfl_total:5.1f}")
    
    print(f"\n  共读取 {len(updates)} 条记录")
    
    print("\n" + "="*60)
    print("[2/3] 更新数据库")
    print("="*60 + "\n")
    
    # 更新ironman_scores表
    for data in updates:
        cursor.execute('''
            INSERT OR REPLACE INTO ironman_scores 
            (player_id, sport, regular_rank, regular_points, playoff_points, total_points, is_final)
            VALUES (?, ?, ?, ?, ?, ?, 1)
        ''', (
            data['player_id'],
            data['sport'],
            data['regular_rank'],
            data['regular_points'],
            data['playoff_points'],
            data['total_points']
        ))
        print(f"  ✓ {data['sport']:4s} | {data['player_name']:15s} | 总计: {data['total_points']:5.1f}")
    
    conn.commit()
    
    print("\n" + "="*60)
    print("[3/3] 更新排行榜")
    print("="*60 + "\n")
    
    # 重新计算排行榜
    cursor.execute('''
        DELETE FROM ironman_leaderboard
    ''')
    
    cursor.execute('''
        INSERT INTO ironman_leaderboard (
            player_id, 
            player_name, 
            total_score, 
            mlb_points, 
            nfl_points, 
            nhl_points, 
            nba_points,
            completed_sports
        )
        SELECT 
            p.player_id,
            p.player_name,
            COALESCE(SUM(s.total_points), 0) as total_score,
            COALESCE(MAX(CASE WHEN s.sport = 'MLB' THEN s.total_points END), 0) as mlb_points,
            COALESCE(MAX(CASE WHEN s.sport = 'NFL' THEN s.total_points END), 0) as nfl_points,
            COALESCE(MAX(CASE WHEN s.sport = 'NHL' THEN s.total_points END), 0) as nhl_points,
            COALESCE(MAX(CASE WHEN s.sport = 'NBA' THEN s.total_points END), 0) as nba_points,
            COUNT(DISTINCT s.sport) as completed_sports
        FROM players p
        LEFT JOIN ironman_scores s ON p.player_id = s.player_id
        GROUP BY p.player_id, p.player_name
    ''')
    
    cursor.execute('''
        UPDATE ironman_leaderboard
        SET rank = (
            SELECT COUNT(*) + 1
            FROM ironman_leaderboard l2
            WHERE l2.total_score > ironman_leaderboard.total_score
        )
    ''')
    
    conn.commit()
    
    # 显示TOP5
    cursor.execute('''
        SELECT rank, player_name, total_score, mlb_points, nfl_points, nhl_points, nba_points
        FROM ironman_leaderboard
        ORDER BY rank
        LIMIT 5
    ''')
    
    print("【铁人个人赛TOP5】")
    print(f"{'排名':<6} {'玩家':<15} {'总分':<8} {'MLB':<6} {'NFL':<6} {'NHL':<6} {'NBA':<6}")
    print("-"*60)
    for row in cursor.fetchall():
        print(f"#{row[0]:<5} {row[1]:<15} {row[2]:<8.1f} {row[3]:<6.1f} {row[4]:<6.1f} {row[5]:<6.1f} {row[6]:<6.1f}")
    
    # 更新联赛状态
    print("\n" + "="*60)
    print("[4/3] 更新联赛状态")
    print("="*60 + "\n")
    
    cursor.execute("UPDATE leagues SET status = 'completed' WHERE sport IN ('MLB', 'NFL')")
    print("  ✓ MLB状态: completed")
    print("  ✓ NFL状态: completed")
    
    conn.commit()
    conn.close()
    
    print("\n" + "="*60)
    print("✓ 重新导入完成！")
    print("="*60)
    print("\n下一步:")
    print("  1. 重启Flask: python ironman_app.py")
    print("  2. 强制刷新浏览器: Ctrl+F5")
    print("  3. 查看更新后的数据")
    print("\n注意:")
    print("  - MLB和NFL的数据现在完全来自Excel")
    print("  - 今后运行sync脚本时会跳过已结束的项目")
    print("  - 只有NHL和NBA会从Yahoo同步")


if __name__ == '__main__':
    excel_path = '2526铁人个人赛.xlsx' if len(sys.argv) < 2 else sys.argv[1]
    db_path = 'ironman.db' if len(sys.argv) < 3 else sys.argv[2]
    
    reimport_completed_sports(excel_path, db_path)
